#!/usr/bin/env python3
"""Export every scored genome to an auditable spreadsheet.

The GA record is an append-only JSONL event log plus one JSON file per
candidate. That is complete but not reviewable: you cannot sort 4000 lines of
JSONL by profit factor, or see at a glance which genomes were suppressed.

Writes a workbook with one row per evaluation so every genome and its numbers
can be audited as they go:

  Summary          counts, bests, and the objective in force
  All Candidates   every candidate_scored event, newest first
  Above Objective  fully measured genomes clearing the objective
  Champions        promotion history
  Suppressed       better earners refused the title, with the blocking metric

Safe to re-run at any time; it rebuilds from the event log and keeps no state
of its own.
"""
from __future__ import annotations

import argparse
import json
import math
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_STATE = ROOT / "runtime" / "market-evolution"

# Identity, then the objective, then the evidence, then the tie-breakers.
COLUMNS = [
    ("generation", "Gen"),
    ("at", "Timestamp (UTC)"),
    ("genome_id", "Genome"),
    ("status", "Status"),
    ("evaluated_folds", "Folds"),
    ("min_profit_factor", "Profit Factor"),
    ("min_expectancy", "Expectancy"),
    ("min_coverage", "Coverage"),
    ("min_acted_observations", "Acted Obs"),
    ("min_accuracy", "Accuracy"),
    ("min_balanced_accuracy", "Balanced Acc"),
    ("min_mcc", "MCC"),
    ("min_baseline_margin", "Baseline Margin"),
    ("max_ece", "Max ECE"),
    ("max_drawdown", "Max Drawdown"),
    ("fitness", "Fitness"),
    ("learner_kind", "Learner"),
    ("feature_count", "Features"),
    ("evaluation_signature", "Eval Signature"),
]

RATIO_KEYS = {
    "min_profit_factor", "min_expectancy", "min_coverage", "min_accuracy",
    "min_balanced_accuracy", "min_mcc", "min_baseline_margin", "max_ece",
    "max_drawdown",
}

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
GOOD_FILL = PatternFill("solid", fgColor="C6EFCE")
BAD_FILL = PatternFill("solid", fgColor="FFC7CE")
WARN_FILL = PatternFill("solid", fgColor="FFEB9C")


def read_events(state_dir: Path) -> list[dict[str, Any]]:
    path = state_dir / "events.jsonl"
    if not path.is_file():
        return []
    out: list[dict[str, Any]] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            out.append(json.loads(line))
        except ValueError:
            continue
    return out


def genome_details(state_dir: Path, genome_id: str) -> dict[str, Any]:
    """Learner and feature count live in the candidate file, not the event."""
    path = state_dir / "candidates" / f"{genome_id}.json"
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        return {}
    return {
        "learner_kind": payload.get("learner_kind"),
        "feature_count": len(payload.get("features") or []),
    }


def infer_folds(row: dict[str, Any]) -> int:
    """A `screened` status means the full walk-forward ran.

    Recorded explicitly on newer events; inferred for historical rows so old
    evaluations stay comparable instead of blank.
    """
    folds = row.get("evaluated_folds")
    if isinstance(folds, int) and folds > 0:
        return folds
    return 3 if row.get("status") == "screened" else 1


def scored_rows(events: Iterable[dict[str, Any]], state_dir: Path) -> list[dict[str, Any]]:
    cache: dict[str, dict[str, Any]] = {}
    rows: list[dict[str, Any]] = []
    for event in events:
        if event.get("event") != "candidate_scored":
            continue
        summary = event.get("summary") or {}
        genome_id = str(event.get("genome_id") or "")
        if genome_id not in cache:
            cache[genome_id] = genome_details(state_dir, genome_id)
        row: dict[str, Any] = {
            "generation": event.get("generation"),
            "at": str(event.get("at") or "")[:19].replace("T", " "),
            "genome_id": genome_id,
            "status": event.get("status"),
            "fitness": event.get("fitness"),
            "evaluated_folds": event.get("evaluated_folds"),
            "evaluation_signature": event.get("evaluation_signature"),
        }
        row.update({key: summary.get(key) for key, _ in COLUMNS if key in summary})
        row.update(cache[genome_id])
        row["evaluated_folds"] = infer_folds(row)
        rows.append(row)
    return rows


def write_sheet(book: Workbook, title: str, rows: list[dict[str, Any]],
                objective: float) -> None:
    sheet = book.create_sheet(title)
    for index, (_, label) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=1, column=index, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for r, row in enumerate(rows, start=2):
        for c, (key, _) in enumerate(COLUMNS, start=1):
            value = row.get(key)
            if isinstance(value, float) and not math.isfinite(value):
                value = None
            cell = sheet.cell(row=r, column=c, value=value)
            if key in RATIO_KEYS:
                cell.number_format = "0.0000"
            elif key == "fitness":
                cell.number_format = "0.0"
        pf = row.get("min_profit_factor")
        if isinstance(pf, (int, float)):
            target = sheet.cell(row=r, column=6)
            folds = infer_folds(row)
            if pf >= objective and folds >= 3:
                target.fill = GOOD_FILL
            elif pf >= objective:
                target.fill = WARN_FILL
            elif pf < 1.0:
                target.fill = BAD_FILL
    sheet.freeze_panes = "A2"
    if rows:
        sheet.auto_filter.ref = sheet.dimensions
    for index, (_, label) in enumerate(COLUMNS, start=1):
        width = max(len(label) + 2, 12)
        if label in {"Genome", "Eval Signature", "Timestamp (UTC)", "Status"}:
            width = 24
        sheet.column_dimensions[get_column_letter(index)].width = width


def build(state_dir: Path, output: Path, objective: float) -> dict[str, Any]:
    events = read_events(state_dir)
    rows = scored_rows(events, state_dir)
    rows.sort(key=lambda r: (r.get("generation") or 0), reverse=True)

    measured = [r for r in rows if infer_folds(r) >= 3]
    above = sorted(
        (r for r in measured
         if isinstance(r.get("min_profit_factor"), (int, float))
         and r["min_profit_factor"] >= objective),
        key=lambda r: -r["min_profit_factor"],
    )

    champions: list[dict[str, Any]] = []
    suppressed: list[dict[str, Any]] = []
    for event in events:
        name = event.get("event")
        summary = event.get("summary") or {}
        base: dict[str, Any] = {
            "generation": event.get("generation"),
            "at": str(event.get("at") or "")[:19].replace("T", " "),
            "genome_id": event.get("genome_id"),
        }
        base.update({k: summary.get(k) for k, _ in COLUMNS if k in summary})
        if name == "champion_updated":
            champions.append({**base, "status": "champion",
                              "fitness": event.get("fitness")})
        elif name == "higher_profit_candidate_suppressed":
            reasons = "; ".join(
                str(b.get("metric") or b.get("reason"))
                for b in (event.get("blocking_metrics") or [])
            )
            suppressed.append({**base, "status": reasons or "suppressed",
                               "fitness": event.get("candidate_fitness")})
    champions.reverse()
    suppressed.reverse()

    book = Workbook()
    book.remove(book.active)
    summary_sheet = book.create_sheet("Summary")
    write_sheet(book, "All Candidates", rows, objective)
    write_sheet(book, "Above Objective", above, objective)
    write_sheet(book, "Champions", champions, objective)
    write_sheet(book, "Suppressed", suppressed, objective)

    pfs = [r["min_profit_factor"] for r in rows
           if isinstance(r.get("min_profit_factor"), (int, float))]
    measured_pfs = [r["min_profit_factor"] for r in measured
                    if isinstance(r.get("min_profit_factor"), (int, float))]
    facts: list[tuple[str, Any]] = [
        ("Generated (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")),
        ("Objective profit factor", objective),
        ("", ""),
        ("Evaluations recorded", len(rows)),
        ("Fully measured (3 folds)", len(measured)),
        ("Distinct genomes", len({r.get("genome_id") for r in rows})),
        ("", ""),
        ("Best PF (any evidence)", max(pfs) if pfs else None),
        ("Best PF (fully measured)", max(measured_pfs) if measured_pfs else None),
        ("Measured above objective", len(above)),
        ("Measured and profitable (>1.0)", sum(1 for p in measured_pfs if p > 1.0)),
        ("", ""),
        ("Champion promotions", len(champions)),
        ("Better earners suppressed", len(suppressed)),
    ]
    for index, (label, value) in enumerate(facts, start=1):
        summary_sheet.cell(row=index, column=1, value=label).font = Font(bold=bool(label))
        cell = summary_sheet.cell(row=index, column=2, value=value)
        if isinstance(value, float):
            cell.number_format = "0.0000"
    summary_sheet.column_dimensions["A"].width = 34
    summary_sheet.column_dimensions["B"].width = 26

    output.parent.mkdir(parents=True, exist_ok=True)
    book.save(output)
    return {
        "rows": len(rows), "measured": len(measured), "above": len(above),
        "champions": len(champions), "suppressed": len(suppressed),
        "best_measured": max(measured_pfs) if measured_pfs else None,
        "output": str(output),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--state-dir", type=Path, default=DEFAULT_STATE)
    parser.add_argument("--output", type=Path, default=None)
    parser.add_argument("--objective", type=float, default=None,
                        help="profit-factor bar; defaults to the GA objective")
    args = parser.parse_args()

    objective = args.objective
    if objective is None:
        try:
            from scripts.market_evolution_service import OBJECTIVE_PROFIT_FACTOR
            objective = float(OBJECTIVE_PROFIT_FACTOR)
        except Exception:
            objective = 1.10

    output = args.output or (args.state_dir / "genome_audit.xlsx")
    stats = build(args.state_dir, output, objective)
    print(f"wrote {stats['output']}")
    print(f"  evaluations     : {stats['rows']}")
    print(f"  fully measured  : {stats['measured']}")
    print(f"  above objective : {stats['above']}")
    print(f"  champions       : {stats['champions']}")
    print(f"  suppressed      : {stats['suppressed']}")
    if stats["best_measured"] is not None:
        print(f"  best measured PF: {stats['best_measured']:.4f}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
